from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from database import engine, SessionLocal
import models
from sqlalchemy import func
import datetime
import bcrypt
import uuid
import io
import csv
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
from supabase import create_client, Client

# Load environment variables from .env file
load_dotenv()

# Create all tables
models.Base.metadata.create_all(bind=engine)

app = Flask(__name__)
CORS(app, supports_credentials=True, resources={
    r"/api/*": {
        "origins": [
            "https://bizmanager-saas.vercel.app", 
            "http://localhost:5173"
        ],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": "*"
    }
})
app.config["JWT_SECRET_KEY"] = "bizmanager-secret-key-change-in-production"
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = datetime.timedelta(days=7)
jwt = JWTManager(app)

# Supabase Storage configuration
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_KEY")

supabase_client: Client = None
if SUPABASE_URL and SUPABASE_KEY:
    supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}
MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5 MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# SMTP Configuration from environment variables
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')  # Your email address
SMTP_PASS = os.environ.get('SMTP_PASS', '')  # App password (NOT your regular password)
SMTP_FROM_NAME = os.environ.get('SMTP_FROM_NAME', 'BizManager')
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'http://localhost:5173')


def send_reset_email(to_email, reset_token, user_name):
    """Send a password reset email with a secure link."""
    if not SMTP_USER or not SMTP_PASS:
        raise ValueError("SMTP credentials not configured. Set SMTP_USER and SMTP_PASS environment variables.")

    reset_link = f"{FRONTEND_URL}/#/reset-password?token={reset_token}"

    msg = MIMEMultipart('alternative')
    msg['From'] = f"{SMTP_FROM_NAME} <{SMTP_USER}>"
    msg['To'] = to_email
    msg['Subject'] = 'Reset Your BizManager Password'

    # Plain text fallback
    text_content = f"""Hi {user_name},

We received a request to reset your BizManager password.

Click the link below to set a new password:
{reset_link}

This link will expire in 15 minutes.

If you didn't request this, please ignore this email - your password will remain unchanged.

- The BizManager Team"""

    # Styled HTML email
    html_content = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0; padding:0; background-color:#f1f5f9; font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f1f5f9; padding:40px 20px;">
    <tr><td align="center">
      <table width="100%" cellpadding="0" cellspacing="0" style="max-width:480px; background:#ffffff; border-radius:16px; box-shadow:0 4px 24px rgba(0,0,0,0.08); overflow:hidden;">
        <!-- Header -->
        <tr><td style="background:linear-gradient(135deg,#3b82f6,#2563eb); padding:32px 40px; text-align:center;">
          <h1 style="color:#ffffff; margin:0; font-size:22px; font-weight:700;">Password Reset</h1>
        </td></tr>
        <!-- Body -->
        <tr><td style="padding:32px 40px;">
          <p style="color:#334155; font-size:15px; line-height:1.6; margin:0 0 20px;">Hi <strong>{user_name}</strong>,</p>
          <p style="color:#334155; font-size:15px; line-height:1.6; margin:0 0 24px;">We received a request to reset your BizManager password. Click the button below to create a new password:</p>
          <table width="100%" cellpadding="0" cellspacing="0"><tr><td align="center">
            <a href="{reset_link}" style="display:inline-block; background:linear-gradient(135deg,#3b82f6,#2563eb); color:#ffffff; text-decoration:none; padding:14px 40px; border-radius:8px; font-size:15px; font-weight:600; letter-spacing:0.3px;">Reset My Password</a>
          </td></tr></table>
          <p style="color:#94a3b8; font-size:13px; line-height:1.5; margin:24px 0 0; text-align:center;">This link expires in <strong>15 minutes</strong>.</p>
          <hr style="border:none; border-top:1px solid #e2e8f0; margin:24px 0;">
          <p style="color:#94a3b8; font-size:12px; line-height:1.5; margin:0;">If you didn't request a password reset, you can safely ignore this email. Your password will remain unchanged.</p>
          <p style="color:#cbd5e1; font-size:11px; line-height:1.5; margin:16px 0 0; word-break:break-all;">If the button doesn't work, copy and paste this link into your browser:<br>{reset_link}</p>
        </td></tr>
        <!-- Footer -->
        <tr><td style="background:#f8fafc; padding:20px 40px; text-align:center; border-top:1px solid #e2e8f0;">
          <p style="color:#94a3b8; font-size:12px; margin:0;">BizManager - Secure Business Management</p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body>
</html>"""

    msg.attach(MIMEText(text_content, 'plain', 'utf-8'))
    msg.attach(MIMEText(html_content, 'html', 'utf-8'))

    print(f"[EMAIL] Sending reset email to {to_email} via {SMTP_HOST}:{SMTP_PORT}...")
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_USER, to_email, msg.as_string())
    print(f"[EMAIL] Reset email sent successfully to {to_email}")


def get_db():
    db = SessionLocal()
    try:
        return db
    except:
        db.close()
        raise


def get_current_user(db):
    user_id = get_jwt_identity()
    return db.query(models.User).filter(models.User.id == user_id).first()


# ================================
# AUTH ENDPOINTS
# ================================
@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.json
    db = SessionLocal()
    try:
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        name = data.get('name', '').strip()
        business_name = data.get('businessName', '').strip()
        invite_code = data.get('inviteCode', '').strip()

        if not email or not password or not name:
            return jsonify({"error": "Email, password, and name are required"}), 400
        if len(password) < 6:
            return jsonify({"error": "Password must be at least 6 characters"}), 400

        existing = db.query(models.User).filter(models.User.email == email).first()
        if existing:
            return jsonify({"error": "Email already registered"}), 400

        pw_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        if invite_code:
            invite = db.query(models.Invite).filter(
                models.Invite.code == invite_code,
                models.Invite.used == False
            ).first()
            if not invite:
                return jsonify({"error": "Invalid or expired invite code"}), 400
            if invite.email and invite.email.lower() != email:
                return jsonify({"error": "This invite code is for a different email"}), 400

            user = models.User(
                email=email, password_hash=pw_hash, name=name,
                role="employee", business_id=invite.business_id
            )
            invite.used = True
            db.add(user)
            db.commit()
            db.refresh(user)
            business = db.query(models.Business).filter(models.Business.id == user.business_id).first()
        else:
            if not business_name:
                return jsonify({"error": "Business name is required for owner registration"}), 400
            business = models.Business(name=business_name)
            db.add(business)
            db.flush()
            user = models.User(
                email=email, password_hash=pw_hash, name=name,
                role="owner", business_id=business.id
            )
            db.add(user)
            db.commit()
            db.refresh(user)

        token = create_access_token(identity=str(user.id))
        return jsonify({
            "token": token,
            "user": {"id": user.id, "email": user.email, "name": user.name,
                     "role": user.role, "businessName": business.name, "businessId": business.id}
        }), 201
    finally:
        db.close()


@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    db = SessionLocal()
    try:
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')

        user = db.query(models.User).filter(models.User.email == email).first()
        if not user or not bcrypt.checkpw(password.encode('utf-8'), user.password_hash.encode('utf-8')):
            return jsonify({"error": "Invalid email or password"}), 401

        business = db.query(models.Business).filter(models.Business.id == user.business_id).first()
        token = create_access_token(identity=str(user.id))
        return jsonify({
            "token": token,
            "user": {"id": user.id, "email": user.email, "name": user.name,
                     "role": user.role, "businessName": business.name, "businessId": business.id}
        }), 200
    finally:
        db.close()


@app.route('/api/auth/me', methods=['GET'])
@jwt_required()
def get_me():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        if not user:
            return jsonify({"error": "User not found"}), 404
        business = db.query(models.Business).filter(models.Business.id == user.business_id).first()
        return jsonify({
            "id": user.id, "email": user.email, "name": user.name,
            "role": user.role, "businessName": business.name, "businessId": business.id
        }), 200
    finally:
        db.close()


@app.route('/api/auth/invite', methods=['POST'])
@jwt_required()
def create_invite():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        if user.role != 'owner':
            return jsonify({"error": "Only owners can invite employees"}), 403

        data = request.json
        email = data.get('email', '').strip().lower() if data.get('email') else None
        code = uuid.uuid4().hex[:8].upper()

        invite = models.Invite(code=code, email=email, business_id=user.business_id)
        db.add(invite)
        db.commit()
        return jsonify({"code": code, "email": email}), 201
    finally:
        db.close()


@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    data = request.json
    db = SessionLocal()
    try:
        email = data.get('email', '').strip().lower()
        if not email:
            return jsonify({"error": "Email is required"}), 400

        # Always return same message to prevent email enumeration
        success_message = "If an account with that email exists, a password reset link has been sent."

        user = db.query(models.User).filter(models.User.email == email).first()
        if not user:
            return jsonify({"message": success_message}), 200

        # Invalidate any existing unused tokens for this email
        db.query(models.PasswordResetToken).filter(
            models.PasswordResetToken.email == email,
            models.PasswordResetToken.used == False
        ).update({"used": True})

        # Generate a secure token (full UUID for security)
        token = uuid.uuid4().hex
        expires_at = datetime.datetime.utcnow() + datetime.timedelta(minutes=15)

        reset_token = models.PasswordResetToken(
            email=email, token=token, expires_at=expires_at
        )
        db.add(reset_token)
        db.commit()

        # Send reset email
        try:
            send_reset_email(email, token, user.name)
        except ValueError as e:
            return jsonify({"error": str(e)}), 500
        except Exception as e:
            print(f"Failed to send reset email: {e}")
            return jsonify({"error": "Failed to send reset email. Please check SMTP configuration."}), 500

        return jsonify({"message": success_message}), 200
    finally:
        db.close()


@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    data = request.json
    db = SessionLocal()
    try:
        token = data.get('token', '').strip()
        new_password = data.get('newPassword', '')

        if not token or not new_password:
            return jsonify({"error": "Token and new password are required"}), 400
        if len(new_password) < 6:
            return jsonify({"error": "Password must be at least 6 characters"}), 400

        reset_token = db.query(models.PasswordResetToken).filter(
            models.PasswordResetToken.token == token,
            models.PasswordResetToken.used == False
        ).first()

        if not reset_token:
            return jsonify({"error": "Invalid or expired reset link. Please request a new one."}), 400

        if datetime.datetime.utcnow() > reset_token.expires_at:
            reset_token.used = True
            db.commit()
            return jsonify({"error": "This reset link has expired. Please request a new one."}), 400

        user = db.query(models.User).filter(models.User.email == reset_token.email).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Update password
        user.password_hash = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        reset_token.used = True
        db.commit()

        return jsonify({"message": "Password reset successfully! You can now sign in with your new password."}), 200
    finally:
        db.close()


@app.route('/api/users', methods=['GET'])
@jwt_required()
def get_users():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        if user.role != 'owner':
            return jsonify({"error": "Only owners can view team"}), 403
        users = db.query(models.User).filter(models.User.business_id == user.business_id).all()
        return jsonify([{"id": u.id, "email": u.email, "name": u.name, "role": u.role,
                         "createdAt": u.created_at.isoformat() if u.created_at else None} for u in users]), 200
    finally:
        db.close()


# ================================
# PRODUCTS
# ================================
@app.route('/api/products', methods=['GET'])
@jwt_required()
def get_products():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        products = db.query(models.Product).filter(models.Product.business_id == user.business_id).all()
        result = []
        for p in products:
            item = {"id": p.id, "name": p.name, "category": p.category,
                    "defaultSellingPrice": p.defaultSellingPrice,
                    "currentStock": p.currentStock, "unit": p.unit or "pcs",
                    "imageUrl": p.image_url}
            if user.role == 'owner':
                item["purchaseCost"] = p.purchaseCost
            result.append(item)
        return jsonify(result), 200
    finally:
        db.close()


@app.route('/api/products', methods=['POST'])
@jwt_required()
def add_product():
    data = request.json
    db = SessionLocal()
    try:
        user = get_current_user(db)
        product = models.Product(
            business_id=user.business_id,
            name=data['name'], category=data.get('category', ''),
            purchaseCost=float(data.get('purchaseCost', 0)),
            defaultSellingPrice=float(data.get('defaultSellingPrice', 0)),
            currentStock=float(data.get('currentStock', 0)),
            unit=data.get('unit', 'pcs')
        )
        db.add(product)
        db.commit()
        db.refresh(product)
        return jsonify({"id": product.id, "imageUrl": product.image_url}), 201
    finally:
        db.close()


@app.route('/api/products/<int:id>', methods=['PUT'])
@jwt_required()
def update_product(id):
    data = request.json
    db = SessionLocal()
    try:
        user = get_current_user(db)
        product = db.query(models.Product).filter(
            models.Product.id == id, models.Product.business_id == user.business_id).first()
        if not product:
            return jsonify({"error": "Not found"}), 404
        if 'name' in data: product.name = data['name']
        if 'category' in data: product.category = data['category']
        if 'purchaseCost' in data and user.role == 'owner':
            product.purchaseCost = float(data['purchaseCost'])
        if 'defaultSellingPrice' in data: product.defaultSellingPrice = float(data['defaultSellingPrice'])
        if 'currentStock' in data: product.currentStock = float(data['currentStock'])
        if 'unit' in data: product.unit = data['unit']
        db.commit()
        return jsonify({"success": True}), 200
    finally:
        db.close()


@app.route('/api/products/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_product(id):
    db = SessionLocal()
    try:
        user = get_current_user(db)
        product = db.query(models.Product).filter(
            models.Product.id == id, models.Product.business_id == user.business_id).first()
        if not product:
            return jsonify({"error": "Not found"}), 404
        has_orders = db.query(models.OrderItem).filter(models.OrderItem.productId == id).first()
        if has_orders:
            return jsonify({"error": "Cannot delete product because it has past sales records."}), 400
        # Clean up image file if exists
        if product.image_url and supabase_client:
            try:
                # Extract filename from the URL
                filename = product.image_url.split('/')[-1]
                supabase_client.storage.from_("product-images").remove([filename])
            except Exception as e:
                print(f"Failed to delete image from Supabase: {e}")
        db.delete(product)
        db.commit()
        return jsonify({"success": True}), 200
    finally:
        db.close()


@app.route('/api/products/<int:id>/image', methods=['POST', 'OPTIONS'])
@jwt_required()
def upload_product_image(id):
    # This prevents the browser preflight from being blocked by JWT
    if request.method == "OPTIONS":
        return jsonify({"success": True}), 200
    db = SessionLocal()
    try:
        user = get_current_user(db)
        product = db.query(models.Product).filter(
            models.Product.id == id, models.Product.business_id == user.business_id).first()
        if not product:
            return jsonify({"error": "Product not found"}), 404

        if 'image' not in request.files:
            return jsonify({"error": "No image file provided"}), 400

        file = request.files['image']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400

        if not allowed_file(file.filename):
            return jsonify({"error": "File type not allowed. Use PNG, JPG, or WebP."}), 400

        # Check file size
        file.seek(0, 2)
        size = file.tell()
        file.seek(0)
        if size > MAX_IMAGE_SIZE:
            return jsonify({"error": "File too large. Maximum size is 5 MB."}), 400

        if not supabase_client:
            return jsonify({"error": "Supabase Storage is not configured."}), 500

        # Delete old image if exists
        if product.image_url:
            try:
                old_filename = product.image_url.split('/')[-1]
                supabase_client.storage.from_("product-images").remove([old_filename])
            except Exception as e:
                print(f"Failed to delete old image from Supabase: {e}")

        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{product.id}_{int(datetime.datetime.utcnow().timestamp())}.{ext}"
        
        # Upload to Supabase
        file_bytes = file.read()
        res = supabase_client.storage.from_("product-images").upload(
            path=filename,
            file=file_bytes,
            file_options={"content-type": file.content_type}
        )

        # Get public URL
        public_url = supabase_client.storage.from_("product-images").get_public_url(filename)
        
        product.image_url = public_url
        db.commit()

        return jsonify({"imageUrl": product.image_url}), 200
    finally:
        db.close()


@app.route('/api/products/<int:id>/image', methods=['DELETE', 'OPTIONS'])
@jwt_required()
def delete_product_image(id):
    # This prevents the browser preflight from being blocked by JWT
    if request.method == "OPTIONS":
        return jsonify({"success": True}), 200
    db = SessionLocal()
    try:
        user = get_current_user(db)
        product = db.query(models.Product).filter(
            models.Product.id == id, models.Product.business_id == user.business_id).first()
        if not product:
            return jsonify({"error": "Product not found"}), 404

        if product.image_url and supabase_client:
            try:
                image_filename = product.image_url.split('/')[-1]
                supabase_client.storage.from_("product-images").remove([image_filename])
            except Exception as e:
                print(f"Failed to delete image from Supabase: {e}")
            product.image_url = None
            db.commit()

        return jsonify({"success": True}), 200
    finally:
        db.close()

# ================================
# VENDORS & PURCHASES API
# ================================

@app.route('/api/vendors', methods=['GET'])
@jwt_required()
def get_vendors():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        vendors = db.query(models.Vendor).filter(models.Vendor.business_id == user.business_id).all()
        return jsonify([{
            "id": v.id, 
            "name": v.name, 
            "contactPhone": v.contact_phone, 
            "contactEmail": v.contact_email,
            "leadTimeDays": v.lead_time_days
        } for v in vendors]), 200
    finally:
        db.close()

@app.route('/api/vendors', methods=['POST'])
@jwt_required()
def add_vendor():
    data = request.json
    db = SessionLocal()
    try:
        user = get_current_user(db)
        vendor = models.Vendor(
            business_id=user.business_id,
            name=data['name'],
            contact_phone=data.get('contactPhone', ''),
            contact_email=data.get('contactEmail', ''),
            lead_time_days=int(data.get('leadTimeDays', 3))
        )
        db.add(vendor)
        db.commit()
        db.refresh(vendor)
        return jsonify({"id": vendor.id}), 201
    finally:
        db.close()

@app.route('/api/vendors/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_vendor(id):
    db = SessionLocal()
    try:
        user = get_current_user(db)
        if user.role != 'owner':
            return jsonify({"error": "Only owners can delete vendors"}), 403
            
        vendor = db.query(models.Vendor).filter(
            models.Vendor.id == id, 
            models.Vendor.business_id == user.business_id
        ).first()
        
        if not vendor:
            return jsonify({"error": "Not found"}), 404
            
        db.delete(vendor)
        db.commit()
        return jsonify({"success": True}), 200
    finally:
        db.close()

@app.route('/api/vendors/<int:id>/purchases', methods=['GET'])
@jwt_required()
def get_vendor_purchases(id):
    db = SessionLocal()
    try:
        user = get_current_user(db)
        vendor = db.query(models.Vendor).filter(
            models.Vendor.id == id,
            models.Vendor.business_id == user.business_id
        ).first()
        
        if not vendor:
            return jsonify({"error": "Vendor not found"}), 404
            
        purchases = db.query(models.Purchase).filter(
            models.Purchase.vendor_id == id,
            models.Purchase.business_id == user.business_id
        ).order_by(models.Purchase.date_received.desc()).all()
        
        result = []
        for p in purchases:
            items = db.query(models.PurchaseItem).filter(models.PurchaseItem.purchase_id == p.id).all()
            item_details = []
            for i in items:
                product = db.query(models.Product).filter(models.Product.id == i.product_id).first()
                item_details.append({
                    "productName": product.name if product else "Unknown Product",
                    "quantity": i.quantity_received,
                    "unitCost": i.unit_cost
                })
            
            result.append({
                "id": p.id,
                "date": p.date_received.isoformat(),
                "invoiceNumber": p.invoice_number,
                "totalCost": p.total_cost,
                "items": item_details
            })
            
        return jsonify(result), 200
    finally:
        db.close()

@app.route('/api/purchases', methods=['POST'])
@jwt_required()
def receive_stock():
    data = request.json
    db = SessionLocal()
    try:
        user = get_current_user(db)
        purchase = models.Purchase(
            business_id=user.business_id,
            vendor_id=data.get('vendorId'),
            invoice_number=data.get('invoiceNumber', ''),
            total_cost=float(data.get('totalCost', 0)),
            notes=data.get('notes', '')
        )
        db.add(purchase)
        db.flush()

        items = data.get('items', [])
        for item in items:
            product = db.query(models.Product).filter(
                models.Product.id == item['productId'],
                models.Product.business_id == user.business_id
            ).first()
            
            if not product:
                continue

            qty = float(item['quantity'])
            cost = float(item['unitCost'])

            purchase_item = models.PurchaseItem(
                purchase_id=purchase.id,
                product_id=product.id,
                quantity_received=qty,
                unit_cost=cost
            )
            db.add(purchase_item)
            
            product.currentStock += qty
            product.purchaseCost = cost 

        db.commit()
        return jsonify({"success": True, "purchaseId": purchase.id}), 200
    finally:
        db.close()


# ================================
# CUSTOMERS & PAYMENTS
# ================================
@app.route('/api/customers', methods=['GET'])
@jwt_required()
def get_customers():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        customers = db.query(models.Customer).filter(models.Customer.business_id == user.business_id).all()
        return jsonify([{"id": c.id, "name": c.name, "phone": c.phone, "address": c.address,
                         "totalOutstandingBalance": c.totalOutstandingBalance} for c in customers]), 200
    finally:
        db.close()


@app.route('/api/customers', methods=['POST'])
@jwt_required()
def add_customer():
    data = request.json
    db = SessionLocal()
    try:
        user = get_current_user(db)
        customer = models.Customer(
            business_id=user.business_id,
            name=data['name'], phone=data.get('phone', ''), address=data.get('address', ''),
            totalOutstandingBalance=float(data.get('totalOutstandingBalance', 0))
        )
        db.add(customer)
        db.commit()
        db.refresh(customer)
        return jsonify({"id": customer.id}), 201
    finally:
        db.close()


@app.route('/api/customers/<int:id>/payment', methods=['POST'])
@jwt_required()
def receive_payment(id):
    data = request.json
    db = SessionLocal()
    try:
        user = get_current_user(db)
        customer = db.query(models.Customer).filter(
            models.Customer.id == id, models.Customer.business_id == user.business_id).first()
        if not customer:
            return jsonify({"error": "Not found"}), 404
        amount = float(data.get('amount', 0))
        customer.totalOutstandingBalance -= amount
        payment = models.PaymentLog(customerId=id, paymentAmount=amount, notes=data.get('notes', ''))
        db.add(payment)
        db.commit()
        return jsonify({"success": True, "newBalance": customer.totalOutstandingBalance}), 200
    finally:
        db.close()


@app.route('/api/customers/<int:id>/orders', methods=['GET'])
@jwt_required()
def get_customer_orders(id):
    db = SessionLocal()
    try:
        user = get_current_user(db)
        customer = db.query(models.Customer).filter(
            models.Customer.id == id, models.Customer.business_id == user.business_id).first()
        if not customer:
            return jsonify({"error": "Not found"}), 404
        orders = db.query(models.Order).filter(
            models.Order.customerId == id, models.Order.business_id == user.business_id
        ).order_by(models.Order.saleDate.desc()).all()
        result = []
        for o in orders:
            items = db.query(models.OrderItem).filter(models.OrderItem.orderId == o.id).all()
            item_details = []
            for i in items:
                product = db.query(models.Product).filter(models.Product.id == i.productId).first()
                item_details.append({
                    "productName": product.name if product else "Unknown",
                    "quantity": i.quantitySold, "price": i.actualSellingPrice
                })
            result.append({
                "id": o.id, "date": o.saleDate.isoformat(),
                "totalValue": o.totalOrderValue, "paidUpfront": o.amountPaidUpfront,
                "balanceAdded": o.balanceAdded, "items": item_details
            })
        return jsonify(result), 200
    finally:
        db.close()


# ================================
# POS & ORDERS
# ================================
@app.route('/api/orders', methods=['POST'])
@jwt_required()
def create_order():
    data = request.json
    db = SessionLocal()
    try:
        user = get_current_user(db)
        customer_id = data.get('customerId')
        amount_paid = float(data.get('amountPaidUpfront', 0))
        items = data.get('items', [])

        total_value = 0.0
        order_items_to_add = []

        for item in items:
            product = db.query(models.Product).filter(
                models.Product.id == item['productId'],
                models.Product.business_id == user.business_id).first()
            if not product:
                return jsonify({"error": "Product not found"}), 400
            qty = float(item['quantity'])
            if product.currentStock < qty:
                return jsonify({"error": f"Insufficient stock for {product.name}"}), 400
            price = float(item['actualSellingPrice'])
            product.currentStock -= qty
            total_value += (price * qty)
            order_items_to_add.append(models.OrderItem(
                productId=product.id, quantitySold=qty,
                actualSellingPrice=price, unitCostAtSale=product.purchaseCost
            ))

        balance_added = max(0, total_value - amount_paid)
        order = models.Order(
            business_id=user.business_id, customerId=customer_id,
            totalOrderValue=total_value, amountPaidUpfront=amount_paid,
            balanceAdded=balance_added
        )
        db.add(order)
        db.flush()

        for oi in order_items_to_add:
            oi.orderId = order.id
            db.add(oi)

        if customer_id and balance_added > 0:
            customer = db.query(models.Customer).filter(
                models.Customer.id == customer_id,
                models.Customer.business_id == user.business_id).first()
            if customer:
                customer.totalOutstandingBalance += balance_added

        db.commit()
        return jsonify({"success": True, "orderId": order.id}), 200
    finally:
        db.close()


# ================================
# DASHBOARD (Owner only)
# ================================
@app.route('/api/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        if user.role != 'owner':
            return jsonify({"error": "Access denied"}), 403

        bid = user.business_id
        products = db.query(models.Product).filter(models.Product.business_id == bid).all()
        total_investment = sum(p.purchaseCost * p.currentStock for p in products)
        low_stock = sum(1 for p in products if p.currentStock <= 5)

        total_revenue = db.query(func.sum(models.Order.totalOrderValue)).filter(
            models.Order.business_id == bid).scalar() or 0.0
        total_credit = db.query(func.sum(models.Customer.totalOutstandingBalance)).filter(
            models.Customer.business_id == bid).scalar() or 0.0

        order_items = db.query(models.OrderItem).join(models.Order).filter(
            models.Order.business_id == bid).all()
        total_profit = sum((oi.actualSellingPrice - oi.unitCostAtSale) * oi.quantitySold for oi in order_items)

        today_start = datetime.datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        todays_orders = db.query(models.Order).filter(
            models.Order.business_id == bid, models.Order.saleDate >= today_start).all()
        today_revenue = sum(o.totalOrderValue for o in todays_orders)
        today_ids = [o.id for o in todays_orders]
        today_items = db.query(models.OrderItem).filter(
            models.OrderItem.orderId.in_(today_ids)).all() if today_ids else []
        today_profit = sum((oi.actualSellingPrice - oi.unitCostAtSale) * oi.quantitySold for oi in today_items)

        return jsonify({
            "totalInvestment": total_investment, "totalRevenue": total_revenue,
            "totalProfit": total_profit, "totalOutstandingCredit": total_credit,
            "lowStockCount": low_stock, "todayRevenue": today_revenue, "todayProfit": today_profit
        }), 200
    finally:
        db.close()


# ================================
# REPORTS (Owner only)
# ================================
@app.route('/api/reports/daily', methods=['GET'])
@jwt_required()
def get_daily_reports():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        if user.role != 'owner':
            return jsonify({"error": "Access denied"}), 403

        bid = user.business_id
        date_from = request.args.get('from')
        date_to = request.args.get('to')

        query = db.query(models.Order).filter(models.Order.business_id == bid)
        if date_from:
            query = query.filter(models.Order.saleDate >= datetime.datetime.fromisoformat(date_from))
        if date_to:
            to_date = datetime.datetime.fromisoformat(date_to) + datetime.timedelta(days=1)
            query = query.filter(models.Order.saleDate < to_date)

        orders = query.all()
        order_items = db.query(models.OrderItem).join(models.Order).filter(
            models.Order.business_id == bid).all()
        oi_by_order = {}
        for oi in order_items:
            oi_by_order.setdefault(oi.orderId, []).append(oi)

        daily_data = {}
        for order in orders:
            d = order.saleDate.strftime('%Y-%m-%d')
            if d not in daily_data:
                daily_data[d] = {"date": d, "revenue": 0, "profit": 0, "orderCount": 0}
            daily_data[d]["revenue"] += order.totalOrderValue
            daily_data[d]["orderCount"] += 1
            items = oi_by_order.get(order.id, [])
            daily_data[d]["profit"] += sum(
                (oi.actualSellingPrice - oi.unitCostAtSale) * oi.quantitySold for oi in items)

        result = sorted(daily_data.values(), key=lambda x: x["date"], reverse=True)
        return jsonify(result), 200
    finally:
        db.close()


@app.route('/api/reports/export', methods=['GET'])
@jwt_required()
def export_report():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        if user.role != 'owner':
            return jsonify({"error": "Access denied"}), 403

        bid = user.business_id
        fmt = request.args.get('format', 'csv')
        date_from = request.args.get('from')
        date_to = request.args.get('to')

        query = db.query(models.Order).filter(models.Order.business_id == bid)
        if date_from:
            query = query.filter(models.Order.saleDate >= datetime.datetime.fromisoformat(date_from))
        if date_to:
            to_date = datetime.datetime.fromisoformat(date_to) + datetime.timedelta(days=1)
            query = query.filter(models.Order.saleDate < to_date)

        orders = query.order_by(models.Order.saleDate.desc()).all()

        rows = []
        for o in orders:
            items = db.query(models.OrderItem).filter(models.OrderItem.orderId == o.id).all()
            customer = db.query(models.Customer).filter(models.Customer.id == o.customerId).first() if o.customerId else None
            for oi in items:
                product = db.query(models.Product).filter(models.Product.id == oi.productId).first()
                rows.append({
                    "Date": o.saleDate.strftime('%Y-%m-%d %H:%M'),
                    "Order ID": o.id,
                    "Customer": customer.name if customer else "Walk-in",
                    "Product": product.name if product else "Unknown",
                    "Quantity": oi.quantitySold,
                    "Selling Price": oi.actualSellingPrice,
                    "Cost Price": oi.unitCostAtSale,
                    "Line Total": oi.actualSellingPrice * oi.quantitySold,
                    "Profit": (oi.actualSellingPrice - oi.unitCostAtSale) * oi.quantitySold,
                    "Amount Paid": o.amountPaidUpfront,
                    "Credit Added": o.balanceAdded
                })

        if fmt == 'xlsx':
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()

            # Sales sheet
            ws = wb.active
            ws.title = "Sales Details"
            headers = ["Date", "Order ID", "Customer", "Product", "Quantity",
                        "Selling Price", "Cost Price", "Line Total", "Profit", "Amount Paid", "Credit Added"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for row_idx, row in enumerate(rows, 2):
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col, value=row[h])

            # Inventory sheet
            ws2 = wb.create_sheet("Inventory Snapshot")
            inv_headers = ["Product", "Category", "Purchase Cost", "Selling Price", "Stock", "Unit", "Total Investment"]
            for col, h in enumerate(inv_headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
            products = db.query(models.Product).filter(models.Product.business_id == bid).all()
            for i, p in enumerate(products, 2):
                ws2.cell(row=i, column=1, value=p.name)
                ws2.cell(row=i, column=2, value=p.category)
                ws2.cell(row=i, column=3, value=p.purchaseCost)
                ws2.cell(row=i, column=4, value=p.defaultSellingPrice)
                ws2.cell(row=i, column=5, value=p.currentStock)
                ws2.cell(row=i, column=6, value=p.unit or "pcs")
                ws2.cell(row=i, column=7, value=p.purchaseCost * p.currentStock)

            # Summary sheet
            ws3 = wb.create_sheet("Summary")
            total_rev = sum(r["Line Total"] for r in rows)
            total_prof = sum(r["Profit"] for r in rows)
            ws3.cell(row=1, column=1, value="Metric").font = Font(bold=True)
            ws3.cell(row=1, column=2, value="Value").font = Font(bold=True)
            summary = [("Report Period", f"{date_from or 'All'} to {date_to or 'All'}"),
                       ("Total Orders", len(set(r["Order ID"] for r in rows))),
                       ("Total Revenue", total_rev), ("Total Profit", total_prof),
                       ("Total Items Sold", sum(r["Quantity"] for r in rows))]
            for i, (k, v) in enumerate(summary, 2):
                ws3.cell(row=i, column=1, value=k)
                ws3.cell(row=i, column=2, value=v)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            fname = f"BizManager_Report_{date_from or 'all'}_to_{date_to or 'all'}.xlsx"
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=fname)
        else:
            output = io.StringIO()
            if rows:
                writer = csv.DictWriter(output, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)
            fname = f"BizManager_Report_{date_from or 'all'}_to_{date_to or 'all'}.csv"
            return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                             as_attachment=True, download_name=fname)
    finally:
        db.close()


# ================================
# FACTORY RESET (Owner only)
# ================================
@app.route('/api/factory-reset', methods=['POST'])
@jwt_required()
def factory_reset():
    db = SessionLocal()
    try:
        user = get_current_user(db)
        if user.role != 'owner':
            return jsonify({"error": "Access denied"}), 403
        bid = user.business_id
        db.query(models.OrderItem).filter(
            models.OrderItem.orderId.in_(
                db.query(models.Order.id).filter(models.Order.business_id == bid)
            )).delete(synchronize_session=False)
        db.query(models.PaymentLog).filter(
            models.PaymentLog.customerId.in_(
                db.query(models.Customer.id).filter(models.Customer.business_id == bid)
            )).delete(synchronize_session=False)
        db.query(models.Order).filter(models.Order.business_id == bid).delete()
        db.query(models.Customer).filter(models.Customer.business_id == bid).delete()
        db.query(models.Product).filter(models.Product.business_id == bid).delete()
        db.commit()
        return jsonify({"success": True}), 200
    finally:
        db.close()


@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({"status": "healthy"}), 200


if __name__ == '__main__':
    app.run(debug=True, port=5000)
